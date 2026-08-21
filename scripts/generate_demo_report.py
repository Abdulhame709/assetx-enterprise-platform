from __future__ import annotations

from datetime import date
from html import escape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "demo"
OUT.mkdir(parents=True, exist_ok=True)

ASSETS = [
    ["AST-0001", "حاسوب محمول Dell Latitude", "تقنية المعلومات", "المقر الرئيسي / الإدارة", "نشط", "أحمد القحطاني", 1, "2025-01-15", 4850.00, "IT-001"],
    ["AST-0002", "حاسوب محمول Lenovo ThinkPad", "تقنية المعلومات", "المقر الرئيسي / المالية", "نشط", "سارة العتيبي", 1, "2025-02-02", 5120.00, "IT-002"],
    ["AST-0003", "طابعة ليزر متعددة الوظائف", "تقنية المعلومات", "المقر الرئيسي / الموارد البشرية", "قيد الصيانة", "—", 1, "2024-07-20", 7350.00, "IT-003"],
    ["AST-0004", "خادم ملفات مركزي", "تقنية المعلومات", "مركز البيانات", "نشط", "—", 1, "2023-11-01", 28600.00, "IT-004"],
    ["AST-0005", "مكتب إداري", "الأثاث والتجهيزات", "المقر الرئيسي / الإدارة", "نشط", "خالد المطيري", 1, "2022-03-18", 1850.00, "FUR-001"],
    ["AST-0006", "كرسي عمل مريح", "الأثاث والتجهيزات", "المقر الرئيسي / الإدارة", "نشط", "خالد المطيري", 1, "2022-03-18", 620.00, "FUR-002"],
    ["AST-0007", "جهاز عرض Projector", "الأجهزة السمعية والبصرية", "المقر الرئيسي / التدريب", "نشط", "منى الحربي", 1, "2024-01-10", 4300.00, "AV-001"],
    ["AST-0008", "شاشة عرض تفاعلية", "الأجهزة السمعية والبصرية", "الفرع الغربي / التدريب", "نشط", "ناصر الزهراني", 1, "2024-05-05", 15800.00, "AV-002"],
    ["AST-0009", "سيارة خدمة ميدانية", "المركبات", "الفرع الغربي / المواقف", "نشط", "فهد الشهري", 1, "2023-09-12", 98500.00, "VEH-001"],
    ["AST-0010", "سيارة نقل صغيرة", "المركبات", "المستودع المركزي", "قيد الصيانة", "—", 1, "2022-10-28", 74500.00, "VEH-002"],
    ["AST-0011", "جهاز قراءة باركود", "المخزون والمستودعات", "المستودع المركزي", "نشط", "ريم الغامدي", 2, "2025-03-01", 1280.00, "WH-001"],
    ["AST-0012", "ماسح ضوئي للمستندات", "تقنية المعلومات", "الفرع الشرقي / الإدارة", "نشط", "يوسف العبدالله", 1, "2024-08-19", 2750.00, "IT-005"],
    ["AST-0013", "خزانة ملفات معدنية", "الأثاث والتجهيزات", "الفرع الشرقي / الأرشيف", "نشط", "—", 3, "2021-06-10", 2400.00, "FUR-003"],
    ["AST-0014", "جهاز لوحي ميداني", "تقنية المعلومات", "المقر الرئيسي / العمليات", "مفقود", "محمود الشريف", 1, "2025-04-22", 2350.00, "IT-006"],
    ["AST-0015", "جهاز اتصال لاسلكي", "الأجهزة الميدانية", "الفرع الغربي / العمليات", "نشط", "عبدالله السبيعي", 6, "2024-02-14", 5400.00, "OPS-001"],
    ["AST-0016", "مولد كهربائي احتياطي", "المرافق", "المستودع المركزي", "نشط", "—", 1, "2021-12-01", 36500.00, "FAC-001"],
    ["AST-0017", "وحدة تكييف متنقلة", "المرافق", "الفرع الشرقي / الخدمات", "مستبعد", "—", 1, "2020-05-30", 3200.00, "FAC-002"],
    ["AST-0018", "كاميرا مراقبة متنقلة", "الأمن والسلامة", "المقر الرئيسي / الأمن", "نشط", "عمر الحربي", 4, "2025-02-25", 6800.00, "SEC-001"],
]

MOVEMENTS = [
    ["MOV-1001", "AST-0008", "نقل", "الفرع الشرقي / التدريب", "الفرع الغربي / التدريب", "معتمد", "2026-08-03", "توحيد أجهزة التدريب في الفرع الغربي"],
    ["MOV-1002", "AST-0012", "تسليم عهدة", "الفرع الشرقي / الإدارة", "يوسف العبدالله", "معتمد", "2026-08-05", "تسليم الجهاز للمستخدم"],
    ["MOV-1003", "AST-0014", "بلاغ فقدان", "المقر الرئيسي / العمليات", "محمود الشريف", "معلّق", "2026-08-07", "لم يظهر الجهاز أثناء الجرد الميداني"],
    ["MOV-1004", "AST-0010", "إرجاع من الصيانة", "المستودع المركزي", "الورشة المعتمدة", "معلّق", "2026-08-08", "بانتظار تقرير الفحص النهائي"],
    ["MOV-1005", "AST-0003", "إرسال للصيانة", "المقر الرئيسي / الموارد البشرية", "الورشة المعتمدة", "معتمد", "2026-08-09", "عطل في وحدة الطباعة"],
    ["MOV-1006", "AST-0009", "نقل", "الفرع الغربي / المواقف", "المقر الرئيسي / المواقف", "معلّق", "2026-08-10", "تغطية احتياج العمليات المركزية"],
    ["MOV-1007", "AST-0017", "استبعاد", "الفرع الشرقي / الخدمات", "—", "مرفوض", "2026-08-11", "لم تكتمل مرفقات الإتلاف"],
    ["MOV-1008", "AST-0011", "تسليم عهدة", "المستودع المركزي", "ريم الغامدي", "معتمد", "2026-08-12", "تسليم قارئي باركود لفريق المستودع"],
]

INVENTORY = [
    ["INV-3001", "AST-0001", "حاسوب محمول Dell Latitude", "المقر الرئيسي / الإدارة", "المقر الرئيسي / الإدارة", 1, 1, "مطابق", "2026-08-12", "تم التحقق من الرقم التسلسلي"],
    ["INV-3002", "AST-0003", "طابعة ليزر متعددة الوظائف", "المقر الرئيسي / الموارد البشرية", "الورشة المعتمدة", 1, 1, "منقول", "2026-08-12", "يوجد طلب حركة صيانة معتمد"],
    ["INV-3003", "AST-0008", "شاشة عرض تفاعلية", "الفرع الشرقي / التدريب", "الفرع الغربي / التدريب", 1, 1, "منقول", "2026-08-12", "تم إنشاء طلب نقل للمراجعة"],
    ["INV-3004", "AST-0014", "جهاز لوحي ميداني", "المقر الرئيسي / العمليات", "—", 1, 0, "مفقود", "2026-08-12", "تم إنشاء بلاغ فقدان معلّق"],
    ["INV-3005", "AST-0011", "جهاز قراءة باركود", "المستودع المركزي", "المستودع المركزي", 2, 1, "فرق كمية", "2026-08-12", "تم العثور على جهاز واحد من أصل اثنين"],
    ["INV-3006", "AST-0012", "ماسح ضوئي للمستندات", "الفرع الشرقي / الإدارة", "الفرع الشرقي / الإدارة", 1, 1, "مطابق", "2026-08-12", "تم التحقق"],
    ["INV-3007", "AST-0007", "جهاز عرض Projector", "المقر الرئيسي / التدريب", "المقر الرئيسي / التدريب", 1, 1, "مطابق", "2026-08-13", "تم التحقق"],
    ["INV-3008", "AST-0009", "سيارة خدمة ميدانية", "الفرع الغربي / المواقف", "المقر الرئيسي / المواقف", 1, 1, "منقول", "2026-08-13", "في عهدة العمليات المركزية"],
    ["INV-3009", "AST-0015", "جهاز اتصال لاسلكي", "الفرع الغربي / العمليات", "الفرع الغربي / العمليات", 6, 6, "مطابق", "2026-08-13", "تمت مطابقة الكمية"],
    ["INV-3010", "AST-0018", "كاميرا مراقبة متنقلة", "المقر الرئيسي / الأمن", "المقر الرئيسي / الأمن", 4, 4, "مطابق", "2026-08-13", "تمت مطابقة الكمية"],
    ["INV-3011", "AST-0010", "سيارة نقل صغيرة", "المستودع المركزي", "الورشة المعتمدة", 1, 1, "منقول", "2026-08-13", "بانتظار إغلاق أمر الصيانة"],
    ["INV-3012", "AST-0004", "خادم ملفات مركزي", "مركز البيانات", "مركز البيانات", 1, 1, "مطابق", "2026-08-13", "تم التحقق من الرقم التسلسلي"],
]

STATUS_COUNTS = [("نشط", 14), ("قيد الصيانة", 2), ("مفقود", 1), ("مستبعد", 1)]
CATEGORY_COUNTS = [("تقنية المعلومات", 6), ("الأثاث والتجهيزات", 4), ("المركبات", 2), ("الأجهزة الميدانية", 1), ("المرافق", 2), ("أخرى", 3)]
LOCATION_COUNTS = [("المقر الرئيسي", 9), ("الفرع الغربي", 4), ("الفرع الشرقي", 3), ("المستودع المركزي", 2)]

BLUE = "1F4E79"
LIGHT_BLUE = "D6E3F0"
PALE_BLUE = "EEF5FB"
GREEN = "E8F5E9"
ORANGE = "FFF3E0"
RED = "FFEBEE"
GRAY = "5B6573"
BORDER = Side(style="thin", color="D7DEE7")


def money(value: float) -> str:
    return f"{value:,.2f} ر.س"


def make_workbook(path: Path) -> None:
    wb = Workbook()
    overview = wb.active
    overview.title = "الملخص"
    assets_ws = wb.create_sheet("سجل الأصول")
    movements_ws = wb.create_sheet("الحركات")
    inventory_ws = wb.create_sheet("نتائج الجرد")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.rightToLeft = True
        ws.freeze_panes = "B5"
        ws.column_dimensions["A"].width = 3

    total_value = sum(row[8] * row[6] for row in ASSETS)
    total_expected = sum(row[5] for row in INVENTORY)
    total_actual = sum(row[6] for row in INVENTORY)
    matched = sum(1 for row in INVENTORY if row[7] == "مطابق")
    discrepancies = len(INVENTORY) - matched
    completion = matched / len(INVENTORY)

    overview.merge_cells("B2:H2")
    overview["B2"] = "تقرير تشغيلي تجريبي لإدارة الأصول"
    overview["B2"].font = Font(name="Arial", size=18, bold=True, color=BLUE)
    overview["B2"].alignment = Alignment(horizontal="right")
    overview.merge_cells("B3:H3")
    overview["B3"] = "بيانات مصطنعة للعرض فقط — لا تمثل سجلات تشغيلية حقيقية"
    overview["B3"].font = Font(name="Arial", size=11, italic=True, color="9C3D20")
    overview["B3"].alignment = Alignment(horizontal="right")

    metrics = [
        ("إجمالي الأصول", len(ASSETS), "#,##0", PALE_BLUE),
        ("القيمة الدفترية التقريبية", total_value, '#,##0.00" ر.س"', PALE_BLUE),
        ("الأصول النشطة", STATUS_COUNTS[0][1], "#,##0", GREEN),
        ("الحركات المعلقة", sum(1 for row in MOVEMENTS if row[5] == "معلّق"), "#,##0", ORANGE),
        ("اكتمال الجرد", completion, "0.0%", GREEN),
        ("فروقات الجرد", discrepancies, "#,##0", RED),
    ]
    for idx, (label, value, number_format, fill_color) in enumerate(metrics):
        col = 2 + idx
        overview.cell(5, col, label)
        overview.cell(6, col, value)
        for row in (5, 6):
            cell = overview.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        overview.cell(5, col).font = Font(name="Arial", size=10, bold=True, color=BLUE)
        overview.cell(6, col).font = Font(name="Arial", size=15, bold=True, color=BLUE)
        overview.cell(6, col).number_format = number_format
    overview.row_dimensions[5].height = 30
    overview.row_dimensions[6].height = 34

    overview.merge_cells("B9:D9")
    overview["B9"] = "توزيع حالة الأصول"
    overview["B9"].font = Font(name="Arial", size=13, bold=True, color=BLUE)
    overview.merge_cells("F9:H9")
    overview["F9"] = "توزيع الأصول حسب الموقع"
    overview["F9"].font = Font(name="Arial", size=13, bold=True, color=BLUE)

    overview.append([])
    status_start = 11
    overview.cell(status_start, 2, "الحالة")
    overview.cell(status_start, 3, "عدد الأصول")
    for r, (label, count) in enumerate(STATUS_COUNTS, status_start + 1):
        overview.cell(r, 2, label)
        overview.cell(r, 3, count)
    loc_start = 11
    overview.cell(loc_start, 6, "الموقع")
    overview.cell(loc_start, 7, "عدد الأصول")
    for r, (label, count) in enumerate(LOCATION_COUNTS, loc_start + 1):
        overview.cell(r, 6, label)
        overview.cell(r, 7, count)

    for row in overview.iter_rows(min_row=11, max_row=15, min_col=2, max_col=7):
        for cell in row:
            cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
            cell.alignment = Alignment(horizontal="right", vertical="center")
    for cell in overview[11][1:3] + overview[11][5:7]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    chart = DoughnutChart()
    chart.title = "حالة الأصول"
    chart.add_data(Reference(overview, min_col=3, min_row=11, max_row=15), titles_from_data=True)
    chart.set_categories(Reference(overview, min_col=2, min_row=12, max_row=15))
    chart.height = 7
    chart.width = 10
    overview.add_chart(chart, "B18")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "الأصول حسب الموقع"
    bar.add_data(Reference(overview, min_col=7, min_row=11, max_row=15), titles_from_data=True)
    bar.set_categories(Reference(overview, min_col=6, min_row=12, max_row=15))
    bar.height = 7
    bar.width = 12
    overview.add_chart(bar, "F18")

    overview["B28"] = "ملاحظات العرض"
    overview["B28"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
    overview.merge_cells("B29:H30")
    overview["B29"] = "هذا الملف نموذج مرئي يوضح شكل التقرير المتوقع عند استخدام بيانات المنصة الحية. الأرقام والأسماء والحالات مصطنعة، ولم يتم إدخالها إلى قاعدة البيانات أو استخدامها في أي عملية تشغيلية."
    overview["B29"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    overview["B29"].font = Font(name="Arial", size=10, italic=True, color=GRAY)
    overview["B29"].fill = PatternFill("solid", fgColor="F7F9FC")

    def write_table(ws, title, headers, rows, start_row=2):
        last_col = 1 + len(headers)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=last_col)
        title_cell = ws.cell(start_row, 2, title)
        title_cell.font = Font(name="Arial", size=15, bold=True, color=BLUE)
        title_cell.alignment = Alignment(horizontal="right")
        ws.cell(start_row + 1, 2, "بيانات تجريبية للعرض فقط")
        ws.cell(start_row + 1, 2).font = Font(name="Arial", size=10, italic=True, color="9C3D20")
        header_row = start_row + 3
        for col, header in enumerate(headers, 2):
            cell = ws.cell(header_row, col, header)
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
        for row_index, row in enumerate(rows, header_row + 1):
            for col, value in enumerate(row, 2):
                cell = ws.cell(row_index, col, value)
                cell.font = Font(name="Arial", size=10, color="202938")
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"
                if isinstance(value, str) and value in {"مطابق", "معتمد", "نشط"}:
                    cell.fill = PatternFill("solid", fgColor=GREEN)
                elif isinstance(value, str) and value in {"معلّق", "منقول", "فرق كمية", "قيد الصيانة"}:
                    cell.fill = PatternFill("solid", fgColor=ORANGE)
                elif isinstance(value, str) and value in {"مفقود", "مرفوض", "مستبعد"}:
                    cell.fill = PatternFill("solid", fgColor=RED)
        ws.auto_filter.ref = f"B{header_row}:{get_column_letter(last_col)}{header_row + len(rows)}"
        ws.freeze_panes = f"B{header_row + 1}"
        ws.row_dimensions[header_row].height = 32
        for col in range(2, last_col + 1):
            max_length = max(len(str(ws.cell(r, col).value or "")) for r in range(header_row, header_row + len(rows) + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 4, 14), 34)
        return header_row, header_row + len(rows)

    write_table(assets_ws, "سجل الأصول التجريبي", ["كود الأصل", "اسم الأصل", "التصنيف", "الموقع", "الحالة", "المستلم", "الكمية", "تاريخ الشراء", "قيمة الوحدة", "المرجع"], ASSETS)
    write_table(movements_ws, "سجل الحركات التجريبي", ["رقم الحركة", "كود الأصل", "نوع الحركة", "من", "إلى / المستلم", "الحالة", "التاريخ", "السبب"], MOVEMENTS)
    write_table(inventory_ws, "نتائج دورة الجرد التجريبية", ["رقم النتيجة", "كود الأصل", "اسم الأصل", "الموقع المتوقع", "الموقع الفعلي", "الكمية المتوقعة", "الكمية الفعلية", "النتيجة", "تاريخ العد", "ملاحظات"], INVENTORY)

    for ws in (assets_ws, movements_ws, inventory_ws):
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.oddFooter.center.text = "AssetX — بيانات تجريبية للعرض فقط"

    inventory_ws.conditional_formatting.add("G6:G17", DataBarRule(start_type="min", end_type="max", color=BLUE, showValue=True))
    inventory_ws.conditional_formatting.add("F6:G17", ColorScaleRule(start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50, mid_color="FEF3C7", end_type="max", end_color="DCFCE7"))
    overview.page_setup.orientation = "landscape"
    overview.page_setup.fitToWidth = 1
    overview.sheet_properties.pageSetUpPr.fitToPage = True
    overview.oddFooter.center.text = "AssetX — تقرير تجريبي — لا يمثل بيانات تشغيلية"

    wb.save(path)


def html_table(headers, rows, status_col=None):
    head = "".join(f"<th>{escape(str(header))}</th>" for header in headers)
    body_rows = []
    for row in rows:
        cells = []
        for idx, value in enumerate(row):
            cls = ""
            if status_col is not None and idx == status_col:
                if value in {"نشط", "مطابق", "معتمد"}:
                    cls = "ok"
                elif value in {"معلّق", "منقول", "فرق كمية", "قيد الصيانة"}:
                    cls = "warn"
                elif value in {"مفقود", "مرفوض", "مستبعد"}:
                    cls = "bad"
            cells.append(f"<td class=\"{cls}\">{escape(str(value))}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def make_html(path: Path) -> None:
    total_value = sum(row[8] * row[6] for row in ASSETS)
    matched = sum(1 for row in INVENTORY if row[7] == "مطابق")
    discrepancies = len(INVENTORY) - matched
    completion = matched / len(INVENTORY)
    status_rows = "".join(f"<div class=\"bar-row\"><span>{escape(label)}</span><div class=\"bar\"><i style=\"width:{count / max(v for _, v in STATUS_COUNTS) * 100:.0f}%\"></i></div><b>{count}</b></div>" for label, count in STATUS_COUNTS)
    location_rows = "".join(f"<div class=\"bar-row\"><span>{escape(label)}</span><div class=\"bar\"><i style=\"width:{count / max(v for _, v in LOCATION_COUNTS) * 100:.0f}%\"></i></div><b>{count}</b></div>" for label, count in LOCATION_COUNTS)
    html = f"""<!doctype html>
<html lang=\"ar\" dir=\"rtl\">
<head>
<meta charset=\"utf-8\">
<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
<title>تقرير AssetX التجريبي</title>
<style>
:root {{ --blue:#1f4e79; --ink:#202938; --muted:#687385; --line:#d7dee7; --surface:#f7f9fc; --ok:#e8f5e9; --warn:#fff3e0; --bad:#ffebee; }}
* {{ box-sizing:border-box; }} body {{ margin:0; background:#eef2f7; color:var(--ink); font-family:Tahoma,Arial,sans-serif; line-height:1.6; }}
.page {{ max-width:1200px; margin:24px auto; background:white; padding:34px; box-shadow:0 14px 50px rgba(30,50,80,.12); }}
.header {{ display:flex; justify-content:space-between; gap:24px; border-bottom:3px solid var(--blue); padding-bottom:22px; }}
.header h1 {{ margin:0; color:var(--blue); font-size:28px; }} .header p {{ margin:6px 0 0; color:var(--muted); }}
.demo {{ background:#fff3e0; color:#9c3d20; border:1px solid #f1c18d; padding:12px 16px; border-radius:10px; min-width:260px; font-weight:bold; text-align:center; }}
.metrics {{ display:grid; grid-template-columns:repeat(6,1fr); gap:12px; margin:24px 0; }} .metric {{ border:1px solid var(--line); border-radius:10px; padding:15px 10px; background:var(--surface); text-align:center; }} .metric span {{ display:block; color:var(--muted); font-size:12px; }} .metric b {{ display:block; color:var(--blue); font-size:23px; margin-top:6px; }}
.grid {{ display:grid; grid-template-columns:1fr 1fr; gap:20px; margin:20px 0 30px; }} .panel {{ border:1px solid var(--line); border-radius:10px; padding:18px; }} h2 {{ color:var(--blue); font-size:18px; margin:0 0 14px; border-bottom:1px solid var(--line); padding-bottom:8px; }}
.bar-row {{ display:grid; grid-template-columns:150px 1fr 36px; gap:10px; align-items:center; margin:12px 0; font-size:13px; }} .bar {{ height:14px; border-radius:99px; background:#e8edf3; overflow:hidden; }} .bar i {{ display:block; height:100%; background:var(--blue); border-radius:99px; }}
table {{ width:100%; border-collapse:collapse; margin:12px 0 30px; font-size:12px; }} th {{ background:var(--blue); color:#fff; padding:10px 8px; text-align:right; white-space:nowrap; }} td {{ border:1px solid var(--line); padding:8px; vertical-align:top; }} tbody tr:nth-child(even) {{ background:#fafbfd; }} td.ok {{ background:var(--ok); }} td.warn {{ background:var(--warn); }} td.bad {{ background:var(--bad); color:#9c2027; font-weight:bold; }}
.note {{ background:var(--surface); border-right:4px solid var(--blue); padding:14px 16px; color:var(--muted); margin:18px 0; }} .footer {{ border-top:1px solid var(--line); padding-top:15px; color:var(--muted); font-size:11px; }}
@media (max-width:850px) {{ .page {{ margin:0; padding:18px; }} .header {{ flex-direction:column; }} .metrics {{ grid-template-columns:repeat(2,1fr); }} .grid {{ grid-template-columns:1fr; }} table {{ display:block; overflow-x:auto; white-space:nowrap; }} }}
@media print {{ body {{ background:#fff; }} .page {{ margin:0; box-shadow:none; max-width:none; }} .panel, table {{ break-inside:avoid; }} }}
</style>
</head>
<body>
<main class=\"page\">
<section class=\"header\"><div><h1>تقرير تشغيلي تجريبي لإدارة الأصول</h1><p>الفترة التجريبية: 1–13 أغسطس 2026 · الجهة: شركة النماذج التشغيلية</p></div><div class=\"demo\">بيانات تجريبية للعرض فقط<br><small>لا تمثل سجلات تشغيلية حقيقية</small></div></section>
<section class=\"metrics\">
<div class=\"metric\"><span>إجمالي الأصول</span><b>{len(ASSETS):,}</b></div><div class=\"metric\"><span>القيمة الدفترية التقريبية</span><b>{total_value:,.0f}</b><small>ر.س</small></div><div class=\"metric\"><span>الأصول النشطة</span><b>{STATUS_COUNTS[0][1]}</b></div><div class=\"metric\"><span>الحركات المعلقة</span><b>{sum(1 for row in MOVEMENTS if row[5] == 'معلّق')}</b></div><div class=\"metric\"><span>اكتمال الجرد</span><b>{completion:.1%}</b></div><div class=\"metric\"><span>فروقات الجرد</span><b>{discrepancies}</b></div>
</section>
<div class=\"grid\"><section class=\"panel\"><h2>توزيع حالة الأصول</h2>{status_rows}</section><section class=\"panel\"><h2>توزيع الأصول حسب الموقع</h2>{location_rows}</section></div>
<div class=\"note\"><strong>قراءة تنفيذية:</strong> يظهر النموذج كيف يمكن للتقرير إبراز الأصول النشطة، الحركات التي تحتاج مراجعة، وفروقات الجرد في أعلى الصفحة قبل عرض السجلات التفصيلية. في البيانات التجريبية توجد 6 حالات تحتاج متابعة: أربعة أصول منقولة، وأصل مفقود، وفرق كمية واحد.</div>
<h2>سجل الأصول</h2>{html_table(["الكود", "اسم الأصل", "التصنيف", "الموقع", "الحالة", "المستلم", "الكمية", "قيمة الوحدة"], [[r[0], r[1], r[2], r[3], r[4], r[5], r[6], money(r[8])] for r in ASSETS], 4)}
<h2>سجل الحركات</h2>{html_table(["رقم الحركة", "كود الأصل", "النوع", "من", "إلى / المستلم", "الحالة", "التاريخ", "السبب"], MOVEMENTS, 5)}
<h2>نتائج دورة الجرد</h2>{html_table(["رقم النتيجة", "كود الأصل", "اسم الأصل", "الموقع المتوقع", "الموقع الفعلي", "المتوقع", "الفعلي", "النتيجة", "الملاحظات"], [[r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[9]] for r in INVENTORY], 7)}
<div class=\"footer\">تم إنشاء هذا العرض من بيانات ثابتة مصطنعة بهدف معاينة شكل التقرير فقط. لم يتم حفظ البيانات في قاعدة البيانات المحلية أو السحابية، ولم تُنشأ منها أي حركة أو نتيجة جرد فعلية.</div>
</main></body></html>"""
    path.write_text(html, encoding="utf-8")


if __name__ == "__main__":
    make_workbook(OUT / "assetx-demo-report.xlsx")
    make_html(OUT / "assetx-demo-report-preview.html")
    print(f"Generated: {OUT / 'assetx-demo-report.xlsx'}")
    print(f"Generated: {OUT / 'assetx-demo-report-preview.html'}")
